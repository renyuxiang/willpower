#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""科室bert"""
import os
import copy
import json
import xlrd
from tqdm import tqdm
import time
import random
import numpy as np
import torch
import openpyxl
import torch.nn as nn
from torch.utils.data import Dataset, DataLoader, TensorDataset
from torch.utils.data.sampler import  WeightedRandomSampler
from transformers import BertConfig, BertTokenizer, BertModel, BertForSequenceClassification
from transformers import AdamW, WarmupLinearSchedule
random.seed(100)

MODEL_CLASSES = {
    'bert': (BertConfig, BertForSequenceClassification, BertTokenizer)
    # ,'xlnet': (XLNetConfig, XLNetForSequenceClassification, XLNetTokenizer),
    # 'xlm': (XLMConfig, XLMForSequenceClassification, XLMTokenizer),
    # 'roberta': (RobertaConfig, RobertaForSequenceClassification, RobertaTokenizer),
    # 'distilbert': (DistilBertConfig, DistilBertForSequenceClassification, DistilBertTokenizer)
}

label_num_dict = {line.split('=')[0]: index for index, line in enumerate(open(file='dept.txt', mode='r', encoding='utf-8').readlines())}
print(label_num_dict)
num_label_dict = {v:k for k, v in label_num_dict.items()}
print(num_label_dict)
print('is_availabel')

print(torch.version.cuda)
print(os.environ['CUDA_VISIBLE_DEVICES'])
print(list(range(len(os.environ['CUDA_VISIBLE_DEVICES'].split(',')))))

def get_train_test_data(r_path, sheet_index=0, start_line=0, cols=2):
    data = xlrd.open_workbook(r_path)
    table = data.sheets()[sheet_index]
    # 行数
    row_count = table.nrows
    # 列数
    col_count = table.ncols
    print('rows: ', row_count)
    print('cols: ', col_count)
    result = []
    for row_num in range(start_line, row_count):
        row_data = table.row_values(row_num)
        data = []
        col_real_temp = cols
        if len(row_data) < col_real_temp:
            col_real_temp = len(row_data)
        for col_temp in range(0, col_real_temp):
            data_temp = row_data[col_temp]
            data.append(data_temp)
        result.append(data)
    print('result len: %s' % len(result))
    return result

def w2excle(data, sheet_name, col=None, file_name=None, head=None, wb=None):
    if not wb:
        wb = openpyxl.Workbook()
    ws = wb.create_sheet(title=sheet_name)

    num = 0
    if head:
        ws.append(head)
        num += 1
    for index, temp in enumerate(data):
        col_real_temp = col
        if len(temp) < col_real_temp:
            col_real_temp = len(temp)
        for col_temp in range(0, col_real_temp):
            try:
                ws.cell(index + num + 1, col_temp + 1, temp[col_temp])
            except Exception as err:
                print(err)
    if file_name:
        wb.save(file_name)

class InputFeatures(object):
    """
    A single set of features of data.

    Args:
        input_ids: Indices of input sequence tokens in the vocabulary.
        attention_mask: Mask to avoid performing attention on padding token indices.
            Mask values selected in ``[0, 1]``:
            Usually  ``1`` for tokens that are NOT MASKED, ``0`` for MASKED (padded) tokens.
        token_type_ids: Segment token indices to indicate first and second portions of the inputs.
        label: Label corresponding to the input
    """

    def __init__(self, input_ids, attention_mask, token_type_ids, label):
        self.input_ids = input_ids
        self.attention_mask = attention_mask
        self.token_type_ids = token_type_ids
        self.label = label

    def __repr__(self):
        return str(self.to_json_string())

    def to_dict(self):
        """Serializes this instance to a Python dictionary."""
        output = copy.deepcopy(self.__dict__)
        return output

    def to_json_string(self):
        """Serializes this instance to a JSON string."""
        return json.dumps(self.to_dict(), indent=2, sort_keys=True) + "\n"

tokenizer = BertTokenizer.from_pretrained('bert-base-chinese')

mask_padding_with_zero=True
pad_token=0
pad_token_segment_id=0
max_len = 128
epochs = 20
lr = 1e-5
min_loss = None
every_iterator_print = 10
threshold = 0.5
weight_decay = 0.0
adam_epsilon=1e-8
warmup_steps = 0
t_total = 60
w_path = 'history/1/'
all_data = None

data_dir = '/home/renyx/work_check/model_data/intention/'
# data_dir = 'E:\\me\\data\\intention\\'

class DeptDataset(Dataset):
    """Dataset wrapping tensors.

    Each sample will be retrieved by indexing tensors along the first dimension.

    Arguments:
        *tensors (Tensor): tensors that have the same size of the first dimension.
    """

    def __init__(self, name, shuffle = False):
        data = load_data(name=name)
        data = [[temp[0], label_num_dict[temp[3]]] for temp in data]
        if shuffle:
            random.shuffle(data)
        self.data = data
        print(self.__len__())

    def __getitem__(self, index):
        temp = self.data[index]
        inputs = tokenizer.encode_plus(temp[0][:max_len - 1], add_special_tokens=True, max_length=max_len)
        input_ids, token_type_ids = inputs["input_ids"], inputs["token_type_ids"]
        attention_mask = [1 if mask_padding_with_zero else 0] * len(input_ids)
        padding_length = max_len - len(input_ids)
        input_ids = input_ids + ([pad_token] * padding_length)
        attention_mask = attention_mask + ([0 if mask_padding_with_zero else 1] * padding_length)
        token_type_ids = token_type_ids + ([pad_token_segment_id] * padding_length)
        return input_ids, attention_mask, token_type_ids, temp[1]

    def __len__(self):
        return len(self.data)

def convert_examples_to_feature(data):
    inputs = tokenizer.encode_plus(data[0][:max_len-1], add_special_tokens=True, max_length=max_len)
    input_ids, token_type_ids = inputs["input_ids"], inputs["token_type_ids"]
    attention_mask = [1 if mask_padding_with_zero else 0] * len(input_ids)
    padding_length = max_len - len(input_ids)
    input_ids = input_ids + ([pad_token] * padding_length)
    attention_mask = attention_mask + ([0 if mask_padding_with_zero else 1] * padding_length)
    token_type_ids = token_type_ids + ([pad_token_segment_id] * padding_length)
    return torch.tensor([input_ids], dtype=torch.long),torch.tensor(
        [attention_mask], dtype=torch.long), torch.tensor(
        [token_type_ids], dtype=torch.long),torch.tensor([data[1]], dtype=torch.long)

class Trainer(object):
    def __init__(self):
        self.device = 'cuda' if torch.cuda.is_available() else 'cpu'
        self.device_id = list(range(len(os.environ['CUDA_VISIBLE_DEVICES'].split(','))))
        print(self.device)
        self.paraller = True
        print('start load config....')
        self.config = BertConfig.from_pretrained('bert-base-chinese', num_labels=45)
        print('start load source_model....')
        self.source_model = BertForSequenceClassification.from_pretrained('bert-base-chinese', config=self.config)
        if self.paraller:
            print('start load model....')
            self.model = nn.DataParallel(self.source_model, device_ids=self.device_id).to(self.device)
            print('load model finish....')
        else:
            self.model = self.source_model.to(self.device)
        self.criterion = nn.CrossEntropyLoss()
        self.optimizer = torch.optim.Adam(self.model.parameters(), lr=lr)
        self.threshold = torch.Tensor([threshold]).to(self.device)
        self.min_loss = None
        self.model_best_name = w_path + 'dept_best.pkl'
        self.model_name = w_path + 'dept.pkl'
        self.epoch_model_name = w_path + 'dept_%s.pkl'

        # no_decay = ['bias', 'LayerNorm.weight']
        # optimizer_grouped_parameters = [
        #     {'params': [p for n, p in self.model.named_parameters() if not any(nd in n for nd in no_decay)],
        #      'weight_decay': weight_decay},
        #     {'params': [p for n, p in self.model.named_parameters() if any(nd in n for nd in no_decay)],
        #      'weight_decay': 0.0}
        # ]
        # optimizer = AdamW(optimizer_grouped_parameters, lr=lr, eps=adam_epsilon)
        # scheduler = WarmupLinearSchedule(optimizer, warmup_steps=warmup_steps, t_total=t_total)

    def train(self, train_loader, val_loader):
        for epoch in range(1, epochs + 1):
            print()
            start = time.time()
            train_result = self.train_step(epoch, train_loader)
            print('epoch: {}, ms:{}s'.format(epoch, int(time.time()-start)))
            val_result= self.val(epoch, val_loader)
            print('epoch: {}, ms:{}s, train_acc : {} / {} = {} %, train_loss: {:.8f}; val_acc : {} / {} = {} %, val_loss: {:.8f}'.format(
                epoch, int(time.time()-start), train_result['correct'], train_result['total'], train_result['acc'], train_result['loss'],
                val_result['correct'],val_result['total'],val_result['acc'],val_result['loss']))
            val_loss = val_result['loss']
            if self.min_loss is None or (self.min_loss and self.min_loss > val_loss):
                self.save(self.model_best_name)
                print('epoch:%s 保存新参数, val loss from %s to %s' % (epoch, self.min_loss, val_loss))
                self.min_loss = val_loss
            if epoch % 3 == 0:
                self.save(self.epoch_model_name % str(epoch))
        self.save(self.model_name)
        print('start t...')
        # self.t(model=self.model, echo='dept_end')
        print('train ok')

    def save(self, model_name):
        if self.paraller:
            torch.save(self.model.module, model_name)
        else:
            torch.save(self.model, model_name)


    def train_step(self, epoch, train_loader):
        device = self.device
        model = self.model
        optimizer = self.optimizer
        # criterion = self.criterion
        # threshold = self.threshold

        model.train()
        total_step = len(train_loader)
        total = 0
        correct = 0
        losses = 0.0
        loss_step = 1e-5
        result = {}
        with tqdm(total=total_step) as pbar:
            iterator_loss = 0.0
            iterator_step = 1e-5
            for i, batch in enumerate(train_loader):
                batch = tuple(t.to(device) for t in batch)
                inputs = {'input_ids': batch[0],
                          'token_type_ids': batch[2],
                          'attention_mask': batch[1],
                          'labels': batch[3]}

                outputs = model(**inputs)
                loss = outputs[0]
                if self.paraller:
                    loss = loss.mean() # mean() to average on multi-gpu parallel training
                iterator_loss += loss.item()
                iterator_step += 1
                losses += loss.item()
                loss_step += 1


                logits = outputs[1]
                preds = torch.argmax(logits.detach().cpu(), dim=1)
                out_label_ids = inputs['labels'].detach().cpu()
                correct += (preds == out_label_ids).sum().item()
                total += out_label_ids.shape[0]

                loss.backward()
                optimizer.step()
                optimizer.zero_grad()
                # model.zero_grad()
                if (i + 1) % every_iterator_print == 0:
                    pbar.set_postfix({'Epoch': '[{}/{}]'.format(epoch, epochs),
                                      'Step': '[{}/{}]'.format(i + 1, total_step),
                                      'loss': '{:.8f}'.format(iterator_loss / iterator_step),
                                      })
                    pbar.update(every_iterator_print)
                    iterator_loss = 0.0
                    iterator_step = 1e-5
        acc = 100 * correct / total
        loss_mean = losses / loss_step
        result.update({'acc': acc, 'loss': loss_mean, 'correct': correct, 'total': total})
        return result


    def val(self, epoch, loader):
        device = self.device
        model = self.model
        model.eval()
        total = 0
        correct = 0
        losses = 0.0
        loss_step = 1e-5
        result = {}
        for i, batch in enumerate(loader):
            batch = tuple(t.to(device) for t in batch)
            inputs = {'input_ids': batch[0],
                      'token_type_ids': batch[2],
                      'attention_mask': batch[1],
                      'labels': batch[3]}

            outputs = model(**inputs)
            loss = outputs[0]
            if self.paraller:
                loss = loss.mean()  # mean() to average on multi-gpu parallel training

            losses += loss.item()
            loss_step += 1
            logits = outputs[1]
            preds = torch.argmax(logits.detach().cpu(), dim=1)
            out_label_ids = inputs['labels'].detach().cpu()
            correct += (preds == out_label_ids).sum().item()
            total += out_label_ids.shape[0]

        acc = 100 * correct / total
        loss_mean = losses / loss_step
        result.update({'acc': acc, 'loss': loss_mean, 'correct': correct, 'total': total})
        return result


    def t(self, model=None, name=None, **kwargs):
        check_model = model
        if name:
            check_model = torch.load(w_path+name)
        check_model.eval()
        device = self.device
        total = 0
        correct = 0
        result = []
        data = load_data(name='val.txt')
        data = [[temp[0], label_num_dict[temp[3]]] for temp in data]

        if all_data is not None:
            data = data[: all_data]
        for i, data_temp in enumerate(data):
            if i % 1000 == 0:
                print(i)
            batch = convert_examples_to_feature(data_temp)
            batch = tuple(t.to(device) for t in batch)
            inputs = {'input_ids': batch[0],
                      'token_type_ids': batch[2],
                      'attention_mask': batch[1]}
            outputs = check_model(**inputs)

            logits = outputs[0]
            preds = torch.argmax(logits.detach().cpu(), dim=1).item()
            out_label_ids = batch[3].detach().cpu().item()

            equal = 0   # 0 - 不相等
            if preds == out_label_ids:
                equal += 1
                correct += 1
            total += 1
            result.append([data_temp[0], num_label_dict[data_temp[1]], out_label_ids, preds, equal])
        acc = 100 * correct / total
        print('acc : {} / {} = {} %'.format(correct, total, acc))
        head = ['问句1', 'y_cn', 'y', 'y_hat', 'equal']
        w2excle(result, sheet_name='test_result', col=6, file_name=w_path + kwargs['echo'] + '.xlsx', head=head)
        print(kwargs['echo'])


def split_data():
    data = {}
    dir = '/data/home/raogj/work/other_need/dept_classify/data/train_data_20200312'
    train_dataset = []
    val_dataset = []
    with open(file=dir + '/train_data_20200312.txt', encoding='utf-8', mode='r') as f:
        for index, line in enumerate(f):
            if index % 1000 == 0:
                print(index)
            rows = line.strip().split('\t')
            label = rows[3]
            data.setdefault(label, []).append(line)
    for key_temp in data.keys():
        data_temp = data[key_temp]
        random.shuffle(data_temp)
        train_dataset.extend(data_temp[: int(len(data_temp) * 0.96)])
        val_dataset.extend(data_temp[int(len(data_temp) * 0.96): ])
    with open(file='/home/renyx/work_check/model_data/dept/train.txt', mode='w', encoding='utf-8') as f:
        f.writelines(train_dataset)

    with open(file='/home/renyx/work_check/model_data/dept/val.txt', mode='w', encoding='utf-8') as f:
        f.writelines(val_dataset)

def load_data(name):
    result = []
    with open(file='/home/renyx/work_check/model_data/dept/' + name, mode='r', encoding='utf-8') as f:
        for index, line in enumerate(f):
            if index % 10000 ==0:
                print(index)
            if all_data and index > all_data:
                break
            rows = line.strip().split('\t')
            result.append([rows[0], rows[1], rows[2], rows[3]])
    return result

def my_collate(batch):
    all_input_ids = []
    all_attention_mask = []
    all_token_type_ids = []
    all_labels = []
    for temp in batch:
        all_input_ids.append(temp[0])
        all_attention_mask.append(temp[1])
        all_token_type_ids.append(temp[2])
        all_labels.append(temp[3])
    return torch.tensor(all_input_ids, dtype=torch.long), torch.tensor(
        all_attention_mask, dtype=torch.long), torch.tensor(
        all_token_type_ids, dtype=torch.long), torch.tensor(all_labels, dtype=torch.long)


def train_model():
    train_dataset = DeptDataset('train.txt', shuffle=True)
    # train_dataset = DeptDataset('val.txt', shuffle=False)
    val_dataset = DeptDataset('val.txt', shuffle=False)
    # train_dataset = convert_examples_to_features('val.txt', shuffle=True)
    # val_dataset = convert_examples_to_features('val.txt', shuffle=False)
    # weights = [4 if int(label) == 1 else 1 for _, _, label in train_dataset.data]
    # sampler = WeightedRandomSampler(weights, num_samples=8000, replacement=True)
    # sampler = WeightedRandomSampler(weights, num_samples=len(train_dataset.data), replacement=True)
    train_loader = DataLoader(train_dataset, batch_size=128, num_workers=5, shuffle=True, collate_fn=my_collate)
    val_loader = DataLoader(val_dataset, batch_size=32, num_workers=5, collate_fn=my_collate)
    global t_total
    t_total = len(train_loader)
    print('init trainer....')
    trainer = Trainer()
    print('start train ....')
    trainer.train(train_loader, val_loader)
    # trainer.t(name='dept.pkl', echo='dept_last')
def check_model():
    trainer = Trainer()
    trainer.t(name='dept.pkl', echo='dept_last')
    trainer.t(name='dept_best.pkl', echo='dept_end')

if __name__ == '__main__':
    print('start ...')
    # load_data('val.txt', shuffle=True)
    train_model()
    # check_model()
    # trainer = Trainer()
    # test_dataset = QuestionsDataset(sheet_index=2, dic=char_dic)
    # trainer.t(test_dataset, name='siamese_best.pkl', echo='siamese_best')
    # trainer.t(test_dataset, name='siamese.pkl', echo='siamese')
    # for i, data in enumerate(train_dataset.data):
    #     train_dataset.__getitem__(i)
    # count = 0
    # for q1, q1_len, q2, q2_len, labels in train_loader:
    #     print(len([temp for temp in labels if temp]))
    #     count += 1
    # print(count)
    print('ok')