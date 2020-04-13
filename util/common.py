#!/usr/bin/env python
# -*- coding: utf-8 -*-
import xlrd
import xlwt
import os
import csv
import openpyxl
import power_config
from sklearn.metrics import recall_score, f1_score, accuracy_score, roc_auc_score, classification_report

def metric(y_pred, y_true, score):
    acc = accuracy_score(y_true, y_pred)  # 分类准确率
    recall = recall_score(y_true, y_pred)    # 分类召回率
    f1 = f1_score(y_true, y_pred)
    auc = roc_auc_score(y_true, score)



class Data2Vector(object):

    @classmethod
    def generate(cls, file_name=power_config.char_dict_path):
        result = cls()
        result.load_dict(file_name=file_name)
        return result

    def load_dict(self, file_name=power_config.char_dict_path):
        self.word2idx = {}
        with open(file_name, 'r', encoding='utf-8') as f:
            for line_temp in f:
                line_temp = line_temp.strip()
                if not line_temp:
                    continue
                try:
                    _word, _label = line_temp.split('\t')
                    self.word2idx[_word] = _label
                except:
                    print(line_temp)

    def get_vector(self, text):
        seqs = []
        # text无值, 返回空list
        if not text:
            return []
        for word_temp in text:
            if self.word2idx.get(word_temp):
                seqs.append(int(self.word2idx[word_temp]))
        return seqs

    def get_word_vector(self, single_word, **kwargs):
        result = kwargs.get('default_result', 0)
        if self.word2idx.get(single_word):
            result = str(self.word2idx[single_word])
        return result


def get_all_files(path):   # 递归获取指定目录下所有文件的绝对路径（非目录）
    dir_list = os.listdir(path)
    result = []
    for i in dir_list:
        sub_dir = os.path.join(path, i)
        if os.path.isdir(sub_dir):
            result.extend(get_all_files(sub_dir))
        else:   # 此时sub_dir是文件的绝对路径
            result.append(sub_dir)
    return result

def get_train_test_data(r_path, sheet_index=0, start_line=0, cols=2):
    data = xlrd.open_workbook(r_path)
    table = data.sheets()[sheet_index]
    # 行数
    row_count = table.nrows
    # 列数
    col_count = table.ncols
    print('rows: ', row_count)
    print('cols: ', col_count)
    result = []
    for row_num in range(start_line, row_count):
        row_data = table.row_values(row_num)
        data = []
        col_real_temp = cols
        if len(row_data) < col_real_temp:
            col_real_temp = len(row_data)
        for col_temp in range(0, col_real_temp):
            data_temp = row_data[col_temp]
            data.append(data_temp)
        result.append(data)
    print('result len: %s' % len(result))
    return result





def get_lxsx_write(f=None):
    if not f:
        f = xlwt.Workbook()
    style0 = xlwt.XFStyle()
    font = xlwt.Font()
    font.name = '宋体'
    font.height = 0x00FF
    style0.font = font
    # sheet0 = f.add_sheet(sheet_name, cell_overwrite_ok=True)
    return f, style0


def w2csv(data, file_name, head=None):
    with open(file=file_name, encoding='utf-8-sig', mode='w', newline='') as f:
        writer = csv.writer(f)
        if head:
            writer.writerow(head)
        writer.writerows(data)

def w2excle(data, sheet_name, col=None, file_name=None, head=None, wb=None):
    if not wb:
        wb = openpyxl.Workbook()
    ws = wb.create_sheet(title=sheet_name)

    num = 0
    if head:
        ws.append(head)
        num += 1
    for index, temp in enumerate(data):
        col_real_temp = col
        if len(temp) < col_real_temp:
            col_real_temp = len(temp)
        for col_temp in range(0, col_real_temp):
            try:
                ws.cell(index + num + 1, col_temp + 1, temp[col_temp])
            except Exception as err:
                print(err)
    if file_name:
        wb.save(file_name)

def w2excle_with_col(data, sheet_name, col, file_name=None, head=None, writer=None, style=None):
    if not writer:
        writer, style = get_lxsx_write()
    sheet_train = writer.add_sheet(sheet_name, cell_overwrite_ok=True)
    num = 0
    if head:
        for head_index, temp in enumerate(head):
            sheet_train.write(num, head_index, temp, style)
        num += 1
    for index, temp in enumerate(data):
        col_real_temp = col
        if len(temp) < col_real_temp:
            col_real_temp = len(temp)
        for col_temp in range(0, col_real_temp):
            try:
                sheet_train.write(index + num, col_temp, temp[col_temp], style)
            except Exception as err:
                print(err)
    if file_name:
        writer.save(file_name)
    # print('save successful')


def list_all_files(rootdir):
    _files = []
    list = os.listdir(rootdir)  # 列出文件夹下所有的目录与文件
    for i in range(0, len(list)):
        path = os.path.join(rootdir, list[i])
        if os.path.isdir(path):
            _files.extend(list_all_files(path))
        if os.path.isfile(path):
            _files.append(path)
    return _files


if __name__ == '__main__':
    result = get_train_test_data('E:\\work\\wy_data\\疾病诊断数据\\xwyz自诊.xls', start_line=0, cols=6)
    print(len(result))
